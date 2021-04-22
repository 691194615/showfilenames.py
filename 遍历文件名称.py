import os
import openpyxl.styles
import openpyxl
files_list=os.listdir()
print(files_list)
wb=openpyxl.Workbook()
ws=wb.active
orange_fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor="FFC125")
for i in range(1,len(files_list)+1):
    ws.cell(1,1).value="文件名称"
    thin=openpyxl.styles.Side(border_style="thin",color="000000")
    border=openpyxl.styles.Border(top=thin,left=thin,bottom=thin,right=thin)
    font1 = openpyxl.styles.Font(name="微软雅黑",size=14,bold=True,color="000000")
    font2 = openpyxl.styles.Font(name="微软雅黑", size=12, bold=False,color="000000")
    ws.cell(1,1).border=border
    ws.cell(1,1).fill=orange_fill
    ws.cell(1,1).font=font1
    ws.cell(i,1).value=files_list[i-1]
    ws.cell(i,1).border=border
    ws.cell(i,1).font=font2
wb.save("遍历的文件名称.xlsx")
